"""
SKYFALL Excel Parser
클라이언트 납품 Excel 파일을 파싱하여 샷별 정보를 반환합니다.

헤더 자동 감지: Shot Code, VFX_Work, Description, Cut Duration, Shot_Colorspace 등
"""
import re
from pathlib import Path
from typing import Optional


# 인식 가능한 헤더 이름 → 내부 필드명
_KNOWN_HEADERS = {
    "shot code": "shot_code", "shot_code": "shot_code", "shotcode": "shot_code",
    "vfx_work": "vfx_work", "vfx work": "vfx_work",
    "description": "description",
    "cut duration": "cut_duration", "cut_duration": "cut_duration",
    "shot_colorspace": "colorspace", "shot colorspace": "colorspace",
    "shot_colorspace": "colorspace",
    "timecode_in": "timecode_in", "timecode in": "timecode_in",
    "timecode_out": "timecode_out", "timecode out": "timecode_out",
    "resolution": "resolution",
    "status": "status",
    "sequence": "sequence",
}


def find_excel(folder: Path) -> Optional[Path]:
    """폴더에서 Excel 파일을 찾습니다. (~$ 임시 파일 제외)"""
    for ext in ("*.xlsx", "*.xls"):
        files = sorted(folder.glob(ext))
        files = [f for f in files if not f.name.startswith("~$")]
        if files:
            return files[0]
    return None


def parse_excel(excel_path: Path) -> list[dict]:
    """
    Excel 파일을 파싱하여 샷별 정보를 반환합니다.

    반환 예시:
    [
        {
            "shot_code": "EP01_S004_0002",
            "vfx_work": "comp\\n1) 기리 옷 가슴쪽 덧방 스티커 리무브",
            "description": "기리 옷 가슴쪽에 로고 가린 덧방 스티커 깨끗하게 지워주세요",
            "cut_duration": "131",
            "colorspace": "Arri4.rec709",
        },
        ...
    ]
    """
    try:
        import openpyxl
    except ImportError as e:
        print(f"  [SKYFALL] openpyxl 필요: pip3 install openpyxl ({e})")
        return []

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    # 헤더 행 자동 감지 (처음 20행 안에서 Shot Code 컬럼 찾기)
    header_row = None
    col_map = {}

    for row in ws.iter_rows(min_row=1, max_row=20):
        for cell in row:
            if cell.value and str(cell.value).strip().lower() in _KNOWN_HEADERS:
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        print(f"  ⚠  Excel 헤더를 찾을 수 없습니다: {excel_path.name}")
        wb.close()
        return []

    # 컬럼 매핑
    for cell in ws[header_row]:
        if cell.value:
            key = str(cell.value).strip().lower()
            if key in _KNOWN_HEADERS:
                col_map[_KNOWN_HEADERS[key]] = cell.column - 1  # 0-based index

    if "shot_code" not in col_map:
        print(f"  ⚠  Shot Code 컬럼을 찾을 수 없습니다: {excel_path.name}")
        wb.close()
        return []

    # 데이터 파싱
    results = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or len(row) <= col_map["shot_code"]:
            continue
        shot_code = row[col_map["shot_code"]]
        if not shot_code or not str(shot_code).strip():
            continue
        shot_code = str(shot_code).strip()

        # 샷코드 형식 검증
        if not re.match(r"^E[P]?\d+_S\d+_\d+$", shot_code, re.I):
            continue

        entry = {"shot_code": shot_code.upper()}
        for field, idx in col_map.items():
            if field != "shot_code" and idx < len(row):
                val = row[idx]
                entry[field] = str(val).strip() if val is not None else ""
        results.append(entry)

    wb.close()
    return results
